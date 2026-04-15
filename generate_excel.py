"""
generate_excel.py
Generates a complete financial tracker Excel workbook.
Sheets: Dashboard, P&L, Cost Analysis, Monthly Trend, Raw Data
Uses openpyxl with full formatting, formulas, and conditional formatting.
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from datetime import date
import os

# ── COLOUR PALETTE ──────────────────────────────────────────────
NAVY   = "1E2761"
BLUE   = "185FA5"
GREEN  = "2D6A35"
RED    = "A32D2D"
AMBER  = "854F0B"
WHITE  = "FFFFFF"
LIGHT  = "F2F4F8"
MID    = "D8DDE8"
BORDER = "C8C3BB"
MUTED  = "9B9690"

# ── DEMO DATA ───────────────────────────────────────────────────
BUSINESS = "Restaurang Björken AB"
ORG      = "559059-3025"
PERIOD   = "March 2026"

MONTHLY_DATA = [
    # month, revenue, staff, food, rent, other
    ("Oct 2025",  412800, 180432, 123840, 54000, 20640),
    ("Nov 2025",  438600, 191268, 126630, 57500, 21930),
    ("Dec 2025",  521400, 222480, 153750, 57500, 26070),
    ("Jan 2026",  389200, 175140, 120600, 61000, 19460),
    ("Feb 2026",  461800, 199536, 140850, 61000, 23090),
    ("Mar 2026",  505900, 224978, 149300, 64500, 28422),
]

TARGETS = {"staff_pct": 40, "food_pct": 31, "rent_pct": 13, "margin_pct": 12}

TOP_EXPENSES = [
    ("Personalkostnader",      "Staff",       224978),
    ("Lokalhyra",              "Rent",         64500),
    ("Sysco Sverige",          "Food & Bev",   43624),
    ("Menigo Foodservice",     "Food & Bev",   35064),
    ("Städ & Renhållning",     "Operations",   18200),
    ("Caspeco AB",             "Software",      2490),
    ("Driftkostnader",         "Operations",    6148),
    ("Marknadsföring",         "Marketing",     5000),
    ("Kontokortsavgifter",     "Finance",       3968),
    ("Försäkringar",           "Operations",    3022),
]

# ── STYLE HELPERS ────────────────────────────────────────────────

def hdr_font(size=11, bold=True, color=WHITE):
    return Font(name="Arial", size=size, bold=bold, color=color)

def body_font(size=10, bold=False, color="000000"):
    return Font(name="Arial", size=size, bold=bold, color=color)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border(sides="all"):
    s = Side(style="thin", color=BORDER)
    n = Side(style=None)
    top = s if "t" in sides or sides == "all" else n
    bot = s if "b" in sides or sides == "all" else n
    lft = s if "l" in sides or sides == "all" else n
    rgt = s if "r" in sides or sides == "all" else n
    return Border(top=top, bottom=bot, left=lft, right=rgt)

def apply_header_row(ws, row, cols, height=22):
    ws.row_dimensions[row].height = height
    for col in cols:
        c = ws.cell(row=row, column=col)
        c.font      = hdr_font()
        c.fill      = fill(NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = thin_border()

def apply_data_cell(ws, row, col, value, fmt=None, bold=False, color="000000", align="left", bg=None):
    c           = ws.cell(row=row, column=col, value=value)
    c.font      = body_font(bold=bold, color=color)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border    = thin_border()
    if bg:
        c.fill = fill(bg)
    if fmt:
        c.number_format = fmt
    return c

def set_col_widths(ws, widths):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

def freeze(ws, cell="A2"):
    ws.freeze_panes = cell


# ── SHEET 1: DASHBOARD ───────────────────────────────────────────

def build_dashboard(wb, data):
    ws = wb.create_sheet("Dashboard", 0)
    ws.sheet_view.showGridLines = False
    ws.tab_color = NAVY

    # Title area
    ws.row_dimensions[1].height = 10
    ws.row_dimensions[2].height = 32
    ws.row_dimensions[3].height = 18
    ws.merge_cells("B2:H2")
    c = ws["B2"]
    c.value      = f"{BUSINESS}  ·  Financial Dashboard  ·  {PERIOD}"
    c.font       = Font(name="Arial", size=16, bold=True, color=NAVY)
    c.alignment  = Alignment(horizontal="left", vertical="center")

    ws["B3"].value     = f"Organisation {ORG}  ·  Generated {date.today().isoformat()}"
    ws["B3"].font      = Font(name="Arial", size=9, color=MUTED)
    ws["B3"].alignment = Alignment(horizontal="left")

    # KPI section header
    ws.row_dimensions[5].height = 14
    ws.merge_cells("B5:H5")
    ws["B5"].value     = "KEY PERFORMANCE INDICATORS — MARCH 2026"
    ws["B5"].font      = Font(name="Arial", size=9, bold=True, color=MUTED)
    ws["B5"].alignment = Alignment(horizontal="left")

    # KPI boxes (row 6-9)
    kpis = [
        ("Revenue",     505900, "kr",  "B6:C9", NAVY,  None,  True),
        ("Net Profit",   38700, "kr",  "D6:E9", GREEN, None,  True),
        ("Margin",        7.65, "%",   "F6:G9", BLUE,  None,  True),
        ("Staff Cost%",  44.5,  "%",   "B11:C14",RED, "Target 40%", False),
        ("Food Cost%",   29.5,  "%",   "D11:E14",GREEN,"Target 31%", False),
        ("Rev Growth",    9.6,  "%",   "F11:G14",NAVY, "vs Feb 2026",False),
    ]

    # KPI layout: col B-C row 6-9, col D-E row 6-9, col F-G row 6-9
    # col B-C row 11-14, col D-E row 11-14, col F-G row 11-14
    kpi_positions = [
        ("B6",  "C9"),  ("D6",  "E9"),  ("F6",  "G9"),
        ("B11", "C14"), ("D11", "E14"), ("F11", "G14"),
    ]
    for idx, (label, value, unit, rng, colour, sub, big) in enumerate(kpis):
        top_left = kpi_positions[idx][0]
        bot_right= kpi_positions[idx][1]
        ws.merge_cells(f"{top_left}:{bot_right}")

        # Parse start row/col from top_left
        col_s = ord(top_left[0]) - ord('A') + 1
        row_s = int(top_left[1:])

        # Fill background for entire block
        for r in range(row_s, row_s + 4):
            for col in range(col_s, col_s + 2):
                cell_obj = ws.cell(row=r, column=col)
                cell_obj.fill = fill(colour)

        # Only write to top-left cell of merged range
        tl = ws.cell(row=row_s, column=col_s)
        tl.value      = f"{label}\n{value:,.0f} {unit}" + (f"\n{sub}" if sub else "")
        tl.font       = Font(name="Arial", size=14 if big else 12, bold=True, color=WHITE)
        tl.alignment  = Alignment(horizontal="center", vertical="center", wrap_text=True)

    set_col_widths(ws, {"A": 2, "B": 16, "C": 16, "D": 16, "E": 16, "F": 16, "G": 16, "H": 2})
    return ws


# ── SHEET 2: P&L ──────────────────────────────────────────────

def build_pl(wb, monthly_data, targets):
    ws = wb.create_sheet("P&L")
    ws.sheet_view.showGridLines = False
    ws.tab_color = BLUE
    freeze(ws, "B3")

    months = [row[0] for row in monthly_data]
    N      = len(months)

    # Title
    ws.merge_cells(f"A1:{get_column_letter(N+2)}1")
    c = ws["A1"]
    c.value      = f"Profit & Loss Statement  ·  {BUSINESS}"
    c.font       = Font(name="Arial", size=13, bold=True, color=NAVY)
    c.alignment  = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    # Column headers row 2
    ws["A2"].value     = "Category"
    ws["A2"].font      = hdr_font(color=WHITE)
    ws["A2"].fill      = fill(NAVY)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A2"].border    = thin_border()

    for i, month in enumerate(months, 2):
        c            = ws.cell(row=2, column=i)
        c.value      = month
        c.font       = hdr_font(color=WHITE)
        c.fill       = fill(NAVY)
        c.alignment  = Alignment(horizontal="right", vertical="center")
        c.border     = thin_border()

    # Target column
    tgt_col = N + 2
    c = ws.cell(row=2, column=tgt_col)
    c.value     = "Target"
    c.font      = hdr_font(color=WHITE)
    c.fill      = fill("2D5A27")
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.border    = thin_border()

    ws.row_dimensions[2].height = 18

    # Data rows
    rows = [
        # (label, data_extractor, fmt, bold, section_header, target_value, bg_when_over)
        ("REVENUE",        None, None, True,  True,  None, None),
        ("Revenue",        lambda r: r[1], '#,##0 "kr"', False, False, None, None),
        ("",               None, None, False, False, None, None),
        ("COSTS",          None, None, True,  True,  None, None),
        ("Staff costs",    lambda r: r[2], '#,##0 "kr"', False, False, None, None),
        ("  % of revenue", lambda r: r[2]/r[1], "0.0%", False, False, targets["staff_pct"]/100, True),
        ("Food & beverage",lambda r: r[3], '#,##0 "kr"', False, False, None, None),
        ("  % of revenue", lambda r: r[3]/r[1], "0.0%", False, False, targets["food_pct"]/100, True),
        ("Rent",           lambda r: r[4], '#,##0 "kr"', False, False, None, None),
        ("  % of revenue", lambda r: r[4]/r[1], "0.0%", False, False, targets["rent_pct"]/100, False),
        ("Other costs",    lambda r: r[5], '#,##0 "kr"', False, False, None, None),
        ("Total costs",    lambda r: r[2]+r[3]+r[4]+r[5], '#,##0 "kr"', True, False, None, None),
        ("",               None, None, False, False, None, None),
        ("NET PROFIT",     None, None, True,  True,  None, None),
        ("Net profit",     lambda r: r[1]-(r[2]+r[3]+r[4]+r[5]), '#,##0 "kr"', True, False, None, None),
        ("Net margin",     lambda r: (r[1]-(r[2]+r[3]+r[4]+r[5]))/r[1], "0.0%", True, False, targets["margin_pct"]/100, False),
    ]

    for row_idx, (label, extractor, fmt, bold, is_section, target, flag_over) in enumerate(rows, 3):
        row_n = row_idx

        # Label cell
        lc           = ws.cell(row=row_n, column=1, value=label)
        lc.font      = Font(name="Arial", size=10, bold=bold, color=NAVY if is_section else "000000")
        lc.alignment = Alignment(horizontal="left", vertical="center")
        if is_section:
            lc.fill = fill(LIGHT)
        lc.border = thin_border()

        for col_idx, data_row in enumerate(monthly_data, 2):
            if extractor is None:
                dc           = ws.cell(row=row_n, column=col_idx, value="")
                dc.fill      = fill(LIGHT) if is_section else fill(WHITE)
                dc.border    = thin_border()
                continue
            val          = extractor(data_row)
            dc           = ws.cell(row=row_n, column=col_idx, value=val)
            dc.font      = Font(name="Arial", size=10, bold=bold,
                                color=GREEN if (not flag_over or val <= (target or 999)) else RED)
            dc.alignment = Alignment(horizontal="right", vertical="center")
            dc.number_format = fmt or "General"
            dc.border    = thin_border()
            if col_idx == N + 1:  # Last actual month — highlight
                dc.fill = fill(LIGHT)

        # Target column
        tc = ws.cell(row=row_n, column=tgt_col)
        if target is not None:
            tc.value          = target
            tc.number_format  = fmt or "General"
            tc.font           = Font(name="Arial", size=10, bold=bold, color=GREEN)
            tc.fill           = fill("EAF2E8")
        tc.border = thin_border()
        tc.alignment = Alignment(horizontal="right")

    set_col_widths(ws, {"A": 22, **{get_column_letter(i+2): 14 for i in range(N)}, get_column_letter(tgt_col): 10})
    ws.row_dimensions[1].height = 26

    # Add conditional formatting — red if % > target
    ws.conditional_formatting.add(
        f"B8:G8",
        CellIsRule(operator="greaterThan", formula=["0.31"], stopIfTrue=False,
                   fill=PatternFill(bgColor="FDECEA"), font=Font(color=RED))
    )

    return ws


# ── SHEET 3: COST ANALYSIS ───────────────────────────────────────

def build_cost_analysis(wb):
    ws = wb.create_sheet("Cost Analysis")
    ws.sheet_view.showGridLines = False
    ws.tab_color = "854F0B"
    freeze(ws, "A3")

    # Header
    ws.merge_cells("A1:E1")
    ws["A1"].value      = f"Top Expense Analysis  ·  March 2026  ·  {BUSINESS}"
    ws["A1"].font       = Font(name="Arial", size=13, bold=True, color=NAVY)
    ws["A1"].alignment  = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    # Column headers
    headers = ["Supplier / Category", "Category", "Amount (kr)", "% of Revenue", "Budget Status"]
    for i, h in enumerate(headers, 1):
        c            = ws.cell(row=2, column=i, value=h)
        c.font       = hdr_font()
        c.fill       = fill(NAVY)
        c.alignment  = Alignment(horizontal="center" if i > 1 else "left", vertical="center")
        c.border     = thin_border()
    ws.row_dimensions[2].height = 18

    revenue = 505900
    total   = sum(e[2] for e in TOP_EXPENSES)

    for i, (supplier, category, amount) in enumerate(TOP_EXPENSES, 3):
        pct = amount / revenue
        bg  = LIGHT if i % 2 == 0 else WHITE
        apply_data_cell(ws, i, 1, supplier, bold=False, bg=bg)
        apply_data_cell(ws, i, 2, category, align="center", bg=bg)
        apply_data_cell(ws, i, 3, amount, fmt='#,##0 "kr"', align="right", bold=i==6, bg=bg)
        apply_data_cell(ws, i, 4, pct, fmt="0.00%", align="right", bg=bg)
        status = "Within budget" if category not in ("Staff",) else ("Over target" if pct > 0.40 else "OK")
        color  = RED if "Over" in status else GREEN
        apply_data_cell(ws, i, 5, status, color=color, align="center", bold=True, bg=bg)

    # Total row
    total_row = len(TOP_EXPENSES) + 3
    apply_data_cell(ws, total_row, 1, "TOTAL SHOWN", bold=True, bg=LIGHT)
    apply_data_cell(ws, total_row, 2, "",  bg=LIGHT)
    apply_data_cell(ws, total_row, 3, total, fmt='#,##0 "kr"', align="right", bold=True, bg=LIGHT)
    apply_data_cell(ws, total_row, 4, total/revenue, fmt="0.00%", align="right", bold=True, bg=LIGHT)
    apply_data_cell(ws, total_row, 5, "", bg=LIGHT)

    set_col_widths(ws, {"A": 28, "B": 16, "C": 16, "D": 14, "E": 16})

    # Bar chart
    chart       = BarChart()
    chart.type  = "bar"
    chart.title = "Expense Breakdown"
    chart.y_axis.title = "Amount (kr)"
    chart.x_axis.title = "Supplier"
    chart.width  = 22
    chart.height = 12
    data = Reference(ws, min_col=3, min_row=2, max_row=len(TOP_EXPENSES)+2)
    cats = Reference(ws, min_col=1, min_row=3, max_row=len(TOP_EXPENSES)+2)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill = "1E2761"
    ws.add_chart(chart, f"G3")

    return ws


# ── SHEET 4: MONTHLY TREND ────────────────────────────────────────

def build_trend(wb, monthly_data):
    ws = wb.create_sheet("Monthly Trend")
    ws.sheet_view.showGridLines = False
    ws.tab_color = GREEN
    freeze(ws, "B2")

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"].value     = f"Monthly Trend  ·  Oct 2025 – Mar 2026  ·  {BUSINESS}"
    ws["A1"].font      = Font(name="Arial", size=13, bold=True, color=NAVY)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    headers = ["Month", "Revenue", "Staff", "Food", "Rent", "Other", "Total Cost", "Net Profit", "Margin %"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=i, value=h)
        c.font      = hdr_font()
        c.fill      = fill(NAVY)
        c.alignment = Alignment(horizontal="right" if i > 1 else "left", vertical="center")
        c.border    = thin_border()
    ws.row_dimensions[2].height = 18

    for row_i, (month, rev, staff, food, rent, other) in enumerate(monthly_data, 3):
        total_cost = staff + food + rent + other
        profit     = rev - total_cost
        margin     = profit / rev

        bg = LIGHT if row_i % 2 == 0 else WHITE
        row_data = [month, rev, staff, food, rent, other, total_cost, profit, margin]
        fmts     = [None, '#,##0 "kr"', '#,##0 "kr"', '#,##0 "kr"', '#,##0 "kr"', '#,##0 "kr"', '#,##0 "kr"', '#,##0 "kr"', "0.0%"]

        for col_i, (val, fmt) in enumerate(zip(row_data, fmts), 1):
            c = ws.cell(row=row_i, column=col_i, value=val)
            c.font      = Font(name="Arial", size=10,
                               bold=(col_i in (2, 8)),
                               color=GREEN if col_i == 8 and profit > 0 else RED if col_i == 8 else "000000")
            c.alignment = Alignment(horizontal="right" if col_i > 1 else "left", vertical="center")
            c.border    = thin_border()
            c.fill      = fill(bg)
            if fmt:
                c.number_format = fmt

    # Formula row — averages
    avg_row = len(monthly_data) + 3
    ws.cell(row=avg_row, column=1, value="AVERAGE").font = Font(name="Arial", size=10, bold=True, color=NAVY)
    for col_i in range(2, 10):
        col_letter = get_column_letter(col_i)
        start_row  = 3
        end_row    = len(monthly_data) + 2
        c = ws.cell(row=avg_row, column=col_i)
        c.value          = f"=AVERAGE({col_letter}{start_row}:{col_letter}{end_row})"
        c.font           = Font(name="Arial", size=10, bold=True, color=NAVY)
        c.fill           = fill(MID)
        c.border         = thin_border()
        c.alignment      = Alignment(horizontal="right")
        c.number_format  = '#,##0 "kr"' if col_i < 9 else "0.0%"

    set_col_widths(ws, {"A": 14, "B": 14, "C": 14, "D": 14, "E": 12, "F": 12, "G": 14, "H": 14, "I": 10})

    # Line chart — revenue and profit trend
    chart        = LineChart()
    chart.title  = "Revenue & Profit Trend"
    chart.y_axis.title = "Amount (kr)"
    chart.x_axis.title = "Month"
    chart.width  = 24
    chart.height = 14

    rev_data  = Reference(ws, min_col=2, min_row=2, max_row=len(monthly_data)+2)
    prof_data = Reference(ws, min_col=8, min_row=2, max_row=len(monthly_data)+2)
    cats      = Reference(ws, min_col=1, min_row=3, max_row=len(monthly_data)+2)

    chart.add_data(rev_data, titles_from_data=True)
    chart.add_data(prof_data, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.line.solidFill = "1E2761"
    chart.series[0].graphicalProperties.line.width     = 20000
    chart.series[1].graphicalProperties.line.solidFill = "2D6A35"
    chart.series[1].graphicalProperties.line.width     = 20000

    ws.add_chart(chart, "K3")
    return ws


# ── SHEET 5: RAW DATA ────────────────────────────────────────────

def build_raw_data(wb, monthly_data):
    ws = wb.create_sheet("Raw Data")
    ws.tab_color = MUTED
    freeze(ws, "A2")

    ws["A1"].value     = f"Raw Data Export  ·  {BUSINESS}  ·  {date.today().isoformat()}"
    ws["A1"].font      = Font(name="Arial", size=11, bold=True, color=NAVY)
    ws.row_dimensions[1].height = 22

    headers = ["Period", "Revenue (kr)", "Staff Cost (kr)", "Food Cost (kr)",
               "Rent (kr)", "Other (kr)", "Total Cost (kr)", "Net Profit (kr)",
               "Staff %", "Food %", "Rent %", "Margin %"]

    for i, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = hdr_font(size=10)
        c.fill = fill(NAVY)
        c.alignment = Alignment(horizontal="right" if i > 1 else "left", vertical="center")

    for row_i, (month, rev, staff, food, rent, other) in enumerate(monthly_data, 3):
        total  = staff + food + rent + other
        profit = rev - total
        row    = [month, rev, staff, food, rent, other, total, profit,
                  staff/rev, food/rev, rent/rev, profit/rev]
        fmts   = [None] + ['#,##0'] * 7 + ['0.0%'] * 4

        for col_i, (val, fmt) in enumerate(zip(row, fmts), 1):
            c = ws.cell(row=row_i, column=col_i, value=val)
            c.font      = Font(name="Arial", size=10)
            c.alignment = Alignment(horizontal="right" if col_i > 1 else "left")
            if fmt: c.number_format = fmt
            if row_i % 2 == 0: c.fill = fill(LIGHT)

    set_col_widths(ws, {get_column_letter(i): 16 for i in range(1, 13)})
    ws.column_dimensions["A"].width = 12
    return ws


# ── MAIN ────────────────────────────────────────────────────────

def generate_excel(output_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    build_dashboard(wb, MONTHLY_DATA)
    build_pl(wb, MONTHLY_DATA, TARGETS)
    build_cost_analysis(wb)
    build_trend(wb, MONTHLY_DATA)
    build_raw_data(wb, MONTHLY_DATA)

    # Workbook properties
    wb.properties.title   = f"{BUSINESS} — Financial Data"
    wb.properties.subject = "Monthly Financial Export"
    wb.properties.creator = "Command Center"

    wb.save(output_path)
    size = os.path.getsize(output_path)
    print(f"Excel saved: {output_path} ({size // 1024}KB)")
    print(f"Sheets: {[ws.title for ws in wb.worksheets]}")


if __name__ == "__main__":
    generate_excel("/home/claude/exports/financial_tracker_march_2026.xlsx")
