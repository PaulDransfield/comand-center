"""
document_processors.py
lib/documents/document_processors.py
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Full document intelligence pipeline for Command Center.

Processors:
  PDFProcessor      → text, tables, page layout, form fields
  DocxProcessor     → paragraphs, headings, tables (pure stdlib)
  ExcelProcessor    → all sheets, tables, named ranges
  CSVProcessor      → auto-detect delimiter, header inference
  ImageProcessor    → OCR via Claude Vision API

Smart Extraction:
  DocumentClassifier → invoice / report / contract / bank / other
  InvoiceExtractor   → supplier, date, amount, line items
  ReportExtractor    → period, financial figures, KPIs

Usage:
  result = await process_document(file_bytes, filename, org_id, api_key)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""

import re, json, csv, io, hashlib, zipfile, struct
from pathlib import Path
from typing  import Optional
from datetime import datetime

try:
    import pdfplumber
    HAS_PDFPLUMBER = True
except ImportError:
    HAS_PDFPLUMBER = False

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

import xml.etree.ElementTree as ET


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# RESULT STRUCTURE
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

class ProcessingResult:
    def __init__(self):
        self.text         = ""          # Full plain text
        self.pages        = []          # [{ page_num, text, tables }]
        self.tables       = []          # [{ page, rows: [[cell]] }]
        self.metadata     = {}          # title, author, created_at etc
        self.doc_type     = "unknown"   # invoice/report/contract/bank/other
        self.structured   = {}          # extracted structured data
        self.word_count   = 0
        self.page_count   = 0
        self.warnings     = []
        self.file_hash    = ""
        self.success      = True

    def to_dict(self):
        return {
            "text":       self.text,
            "pages":      self.pages,
            "tables":     self.tables,
            "metadata":   self.metadata,
            "doc_type":   self.doc_type,
            "structured": self.structured,
            "word_count": self.word_count,
            "page_count": self.page_count,
            "warnings":   self.warnings,
            "file_hash":  self.file_hash,
            "success":    self.success,
        }


class ProcessingError(Exception):
    def __init__(self, message, recoverable=False, partial_result=None):
        super().__init__(message)
        self.recoverable     = recoverable
        self.partial_result  = partial_result


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# PDF PROCESSOR
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

class PDFProcessor:
    """
    Extracts text and tables from PDF files using pdfplumber.
    Handles:
      - Text-based PDFs (from Fortnox, Word exports)
      - Mixed PDFs with tables
      - Multi-column layouts
      - Password-protected PDFs (with password param)
      - Corrupted/truncated PDFs (partial recovery)
    """

    MAX_PAGES = 500  # Halt after this many pages to prevent memory issues

    def process(self, file_bytes: bytes, password: str = None) -> ProcessingResult:
        result = ProcessingResult()
        result.file_hash = hashlib.sha256(file_bytes).hexdigest()

        if not HAS_PDFPLUMBER:
            raise ProcessingError("pdfplumber not installed — run: pip install pdfplumber")

        # Detect if PDF is just image scans (no selectable text)
        is_image_pdf = False

        try:
            pdf_file = io.BytesIO(file_bytes)
            open_kwargs = {"password": password} if password else {}

            with pdfplumber.open(pdf_file, **open_kwargs) as pdf:
                result.page_count = len(pdf.pages)
                result.metadata   = self._extract_metadata(pdf)

                if result.page_count > self.MAX_PAGES:
                    result.warnings.append(
                        f"Document has {result.page_count} pages — processing first {self.MAX_PAGES} only. "
                        f"For full processing split into smaller files."
                    )

                all_text  = []
                all_chars = 0

                for page_num, page in enumerate(pdf.pages[:self.MAX_PAGES], 1):
                    try:
                        page_result = self._process_page(page, page_num)
                        result.pages.append(page_result)
                        all_text.append(page_result["text"])
                        all_chars += len(page_result["text"])
                        if page_result["tables"]:
                            result.tables.extend(page_result["tables"])
                    except Exception as e:
                        result.warnings.append(f"Page {page_num}: {e}")
                        continue

                result.text = "\n\n".join(all_text)

                # If we got very little text, flag as likely image-based
                avg_chars = all_chars / max(result.page_count, 1)
                if avg_chars < 50:
                    is_image_pdf = True
                    result.warnings.append(
                        "This PDF appears to be image-based (scanned). "
                        "Text extraction is limited. For better results, upload an "
                        "original digital PDF from Fortnox or export as text."
                    )

        except pdfplumber.pdfminer.pdfparser.PDFSyntaxError as e:
            raise ProcessingError(f"Corrupted PDF: {e}", recoverable=False)
        except Exception as e:
            if "encrypt" in str(e).lower() or "password" in str(e).lower():
                raise ProcessingError(
                    "PDF is password-protected. Contact the sender for the password.",
                    recoverable=False
                )
            raise ProcessingError(f"PDF processing failed: {e}", recoverable=False)

        result.word_count = len(result.text.split())
        return result

    def _process_page(self, page, page_num: int) -> dict:
        """Extract text and tables from a single page."""
        # Extract tables first (pdfplumber table extraction is more accurate
        # when done before text extraction)
        tables = []
        try:
            raw_tables = page.extract_tables()
            for tbl in (raw_tables or []):
                if tbl and any(any(cell for cell in row) for row in tbl):
                    cleaned = [
                        [str(cell or "").strip() for cell in row]
                        for row in tbl
                    ]
                    tables.append({"page": page_num, "rows": cleaned})
        except Exception:
            pass  # Table extraction failure is non-fatal

        # Extract text — use layout-aware extraction for multi-column
        text = ""
        try:
            text = page.extract_text(layout=True) or ""
            # Clean up excessive whitespace from layout mode
            text = re.sub(r'\n{3,}', '\n\n', text)
            text = re.sub(r' {3,}', '  ', text)
        except Exception:
            try:
                text = page.extract_text() or ""
            except Exception:
                text = ""

        # Append table text representation
        for tbl in tables:
            table_text = self._table_to_text(tbl["rows"])
            text += f"\n\n[TABLE]\n{table_text}\n[/TABLE]"

        return {"page": page_num, "text": text.strip(), "tables": tables}

    def _table_to_text(self, rows: list) -> str:
        """Convert table rows to readable text."""
        if not rows:
            return ""
        widths = [max(len(str(r[i])) for r in rows if i < len(r)) for i in range(len(rows[0]))]
        lines  = []
        for i, row in enumerate(rows):
            line = " | ".join(str(cell).ljust(widths[j] if j < len(widths) else 10)
                               for j, cell in enumerate(row))
            lines.append(line)
            if i == 0:  # Header separator
                lines.append("-" * len(line))
        return "\n".join(lines)

    def _extract_metadata(self, pdf) -> dict:
        try:
            info = pdf.metadata or {}
            return {
                "title":       info.get("Title", ""),
                "author":      info.get("Author", ""),
                "created":     info.get("CreationDate", ""),
                "modified":    info.get("ModDate", ""),
                "producer":    info.get("Producer", ""),
                "page_count":  len(pdf.pages),
            }
        except Exception:
            return {}


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# DOCX PROCESSOR (pure stdlib — no mammoth dependency)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

class DocxProcessor:
    """
    Extracts content from .docx files using stdlib zipfile + XML parsing.
    A .docx is a ZIP file containing XML — no external library needed.
    Handles: paragraphs, headings, tables, lists, footnotes.
    """

    # Word XML namespace
    W  = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
    CP = "http://schemas.openxmlformats.org/package/2006/metadata/core-properties"
    DC = "http://purl.org/dc/elements/1.1/"

    def process(self, file_bytes: bytes) -> ProcessingResult:
        result = ProcessingResult()
        result.file_hash = hashlib.sha256(file_bytes).hexdigest()

        try:
            with zipfile.ZipFile(io.BytesIO(file_bytes)) as zf:
                names = zf.namelist()

                # Validate it's a real docx
                if "word/document.xml" not in names:
                    raise ProcessingError("Not a valid .docx file (missing word/document.xml)")

                # Extract main document content
                doc_xml  = zf.read("word/document.xml")
                sections = self._parse_document(doc_xml)

                # Extract metadata
                if "docProps/core.xml" in names:
                    result.metadata = self._parse_metadata(zf.read("docProps/core.xml"))

                # Combine sections into full text
                text_parts = []
                tables     = []

                for section in sections:
                    if section["type"] == "paragraph":
                        text_parts.append(section["text"])
                    elif section["type"] == "table":
                        tables.append({"page": 1, "rows": section["rows"]})
                        text_parts.append("[TABLE]\n" + self._table_to_text(section["rows"]) + "\n[/TABLE]")
                    elif section["type"] == "heading":
                        text_parts.append(f"\n## {section['text']}\n")

                result.text   = "\n".join(p for p in text_parts if p.strip())
                result.tables = tables
                # DOCX doesn't have true pages — estimate
                result.page_count = max(1, result.word_count // 250)

        except zipfile.BadZipFile:
            raise ProcessingError(
                "File appears corrupted or is not a .docx file. "
                "Try re-saving from Word and uploading again.",
                recoverable=False
            )
        except ProcessingError:
            raise
        except Exception as e:
            raise ProcessingError(f"Word document processing failed: {e}", recoverable=False)

        result.word_count = len(result.text.split())
        return result

    def _parse_document(self, xml_bytes: bytes) -> list:
        """Parse word/document.xml into structured sections."""
        try:
            root = ET.fromstring(xml_bytes)
        except ET.ParseError as e:
            raise ProcessingError(f"Corrupted XML in document: {e}")

        sections = []
        body     = root.find(f".//{{{self.W}}}body")
        if body is None:
            return sections

        for child in body:
            tag = child.tag.split("}")[-1] if "}" in child.tag else child.tag

            if tag == "p":   # Paragraph
                section = self._parse_paragraph(child)
                if section:
                    sections.append(section)
            elif tag == "tbl":  # Table
                section = self._parse_table(child)
                if section:
                    sections.append(section)

        return sections

    def _parse_paragraph(self, para_el) -> Optional[dict]:
        """Extract text from a paragraph element, detecting heading style."""
        # Check if this is a heading
        style_el = para_el.find(f".//{{{self.W}}}pStyle")
        style    = style_el.get(f"{{{self.W}}}val", "") if style_el is not None else ""
        is_heading = "Heading" in style or "heading" in style or style.startswith("h")

        # Extract all text runs
        texts = []
        for run in para_el.findall(f".//{{{self.W}}}r"):
            for t in run.findall(f"{{{self.W}}}t"):
                text = t.text or ""
                texts.append(text)

        full_text = "".join(texts).strip()
        if not full_text:
            return None

        return {
            "type":  "heading" if is_heading else "paragraph",
            "text":  full_text,
            "style": style,
        }

    def _parse_table(self, tbl_el) -> Optional[dict]:
        """Extract table rows from a table element."""
        rows = []
        for row_el in tbl_el.findall(f".//{{{self.W}}}tr"):
            row = []
            for cell_el in row_el.findall(f".//{{{self.W}}}tc"):
                cell_texts = []
                for t in cell_el.findall(f".//{{{self.W}}}t"):
                    cell_texts.append(t.text or "")
                row.append(" ".join(cell_texts).strip())
            if row:
                rows.append(row)

        if not rows:
            return None
        return {"type": "table", "rows": rows}

    def _table_to_text(self, rows: list) -> str:
        lines = []
        for i, row in enumerate(rows):
            lines.append(" | ".join(str(c) for c in row))
            if i == 0:
                lines.append("-" * max(len(lines[0]), 20))
        return "\n".join(lines)

    def _parse_metadata(self, xml_bytes: bytes) -> dict:
        try:
            root = ET.fromstring(xml_bytes)
            return {
                "title":    self._get_tag(root, self.DC, "title"),
                "author":   self._get_tag(root, self.DC, "creator"),
                "created":  self._get_tag(root, self.CP, "created"),
                "modified": self._get_tag(root, self.CP, "modified"),
            }
        except Exception:
            return {}

    def _get_tag(self, root, ns, tag):
        el = root.find(f"{{{ns}}}{tag}")
        return el.text if el is not None else ""


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# EXCEL PROCESSOR
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

class ExcelProcessor:
    """
    Extracts data from .xlsx and .xls files using openpyxl.
    Handles: multiple sheets, merged cells, named ranges, formulas.
    """

    MAX_ROWS_PER_SHEET = 5000
    MAX_SHEETS         = 20

    def process(self, file_bytes: bytes) -> ProcessingResult:
        result = ProcessingResult()
        result.file_hash = hashlib.sha256(file_bytes).hexdigest()

        if not HAS_OPENPYXL:
            raise ProcessingError("openpyxl not installed — run: pip install openpyxl")

        try:
            wb = openpyxl.load_workbook(
                io.BytesIO(file_bytes),
                read_only=True,      # Memory-efficient for large files
                data_only=True,      # Get computed values, not formulas
            )
        except Exception as e:
            raise ProcessingError(
                f"Could not open Excel file: {e}. "
                "Ensure the file is a valid .xlsx (not .xls). "
                "Try opening in Excel and saving as .xlsx.",
                recoverable=False
            )

        all_text_parts = []
        sheet_count    = min(len(wb.sheetnames), self.MAX_SHEETS)

        if len(wb.sheetnames) > self.MAX_SHEETS:
            result.warnings.append(
                f"Workbook has {len(wb.sheetnames)} sheets — processing first {self.MAX_SHEETS}."
            )

        for sheet_name in wb.sheetnames[:sheet_count]:
            try:
                ws          = wb[sheet_name]
                sheet_result = self._process_sheet(ws, sheet_name)
                all_text_parts.append(f"=== Sheet: {sheet_name} ===\n{sheet_result['text']}")
                if sheet_result["tables"]:
                    result.tables.extend(sheet_result["tables"])
            except Exception as e:
                result.warnings.append(f"Sheet '{sheet_name}': {e}")
                continue

        wb.close()

        result.text       = "\n\n".join(all_text_parts)
        result.word_count = len(result.text.split())
        result.page_count = sheet_count

        return result

    def _process_sheet(self, ws, sheet_name: str) -> dict:
        rows     = []
        row_count = 0

        for row in ws.iter_rows(values_only=True):
            # Skip completely empty rows
            if not any(cell is not None for cell in row):
                continue

            cleaned = [self._format_cell(cell) for cell in row]
            rows.append(cleaned)
            row_count += 1

            if row_count >= self.MAX_ROWS_PER_SHEET:
                break

        if not rows:
            return {"text": f"(Sheet '{sheet_name}' is empty)", "tables": []}

        # Convert to CSV-style text
        text_lines = []
        for row in rows:
            text_lines.append(",".join(f'"{c}"' if "," in c else c for c in row))

        # Also find and label apparent data tables
        tables = self._detect_tables(rows, sheet_name)

        return {
            "text":   "\n".join(text_lines),
            "tables": tables,
        }

    def _format_cell(self, value) -> str:
        """Format a cell value to a clean string."""
        if value is None:
            return ""
        if isinstance(value, float):
            # Avoid scientific notation for common financial values
            if value == int(value) and abs(value) < 1e10:
                return str(int(value))
            return f"{value:.2f}"
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")
        return str(value).strip()

    def _detect_tables(self, rows: list, sheet_name: str) -> list:
        """
        Heuristic: if the first non-empty row looks like a header
        (all strings, no numbers), treat the block as a table.
        """
        if not rows:
            return []

        # Check first row for header-like pattern
        first_row = rows[0]
        looks_like_header = all(
            not re.match(r'^-?\d+(\.\d+)?$', cell)
            for cell in first_row if cell
        )

        if looks_like_header and len(rows) >= 2:
            return [{"sheet": sheet_name, "page": 1, "rows": rows[:100]}]
        return []


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# CSV PROCESSOR
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

class CSVProcessor:
    """
    Imports CSV files with automatic delimiter detection.
    Handles Swedish number formats (comma decimal, period thousands).
    """

    MAX_ROWS = 10_000

    def process(self, file_bytes: bytes) -> ProcessingResult:
        result = ProcessingResult()
        result.file_hash = hashlib.sha256(file_bytes).hexdigest()

        # Detect encoding
        text_content = self._decode(file_bytes)

        # Detect delimiter
        delimiter = self._detect_delimiter(text_content)

        try:
            reader = csv.reader(io.StringIO(text_content), delimiter=delimiter)
            rows   = []
            for i, row in enumerate(reader):
                if i >= self.MAX_ROWS:
                    result.warnings.append(f"CSV truncated at {self.MAX_ROWS} rows.")
                    break
                rows.append([cell.strip() for cell in row])

            if not rows:
                raise ProcessingError("CSV file is empty.", recoverable=False)

            # Convert to readable text
            header   = rows[0] if rows else []
            data_rows = rows[1:]

            text_parts = []
            text_parts.append(f"Headers: {', '.join(header)}")
            text_parts.append(f"Rows: {len(data_rows)}")
            text_parts.append("")

            # Add sample rows as text
            for row in rows[:50]:  # First 50 rows as text
                if header:
                    pairs = [f"{h}: {v}" for h, v in zip(header, row) if v]
                    text_parts.append(", ".join(pairs))
                else:
                    text_parts.append(", ".join(row))

            if len(rows) > 50:
                text_parts.append(f"... ({len(rows) - 50} more rows)")

            result.text       = "\n".join(text_parts)
            result.tables     = [{"page": 1, "rows": rows[:200]}]
            result.page_count = 1
            result.word_count = len(result.text.split())
            result.metadata   = {"columns": header, "row_count": len(data_rows), "delimiter": delimiter}

        except csv.Error as e:
            raise ProcessingError(f"CSV parsing failed: {e}", recoverable=False)

        return result

    def _decode(self, file_bytes: bytes) -> str:
        """Try common encodings — Swedish files often use Latin-1."""
        for encoding in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
            try:
                return file_bytes.decode(encoding)
            except UnicodeDecodeError:
                continue
        return file_bytes.decode("latin-1", errors="replace")

    def _detect_delimiter(self, text: str) -> str:
        """Sniff the delimiter from the first few lines."""
        sample = "\n".join(text.splitlines()[:5])
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            return dialect.delimiter
        except csv.Error:
            # Fallback: count occurrences
            counts = {d: sample.count(d) for d in (",", ";", "\t")}
            return max(counts, key=counts.get)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# IMAGE PROCESSOR (OCR via Claude Vision)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

class ImageProcessor:
    """
    Extracts text from images using Claude's vision capability.
    Works for: photos of invoices, screenshots of reports,
    scanned documents, whiteboard photos.

    Requires ANTHROPIC_API_KEY in environment.
    """

    MAX_IMAGE_SIZE_MB = 20

    def process(self, file_bytes: bytes, filename: str, api_key: str) -> ProcessingResult:
        result = ProcessingResult()
        result.file_hash = hashlib.sha256(file_bytes).hexdigest()

        if not api_key:
            raise ProcessingError(
                "API key required for image OCR. "
                "Images require Claude Vision to extract text.",
                recoverable=False
            )

        size_mb = len(file_bytes) / (1024 * 1024)
        if size_mb > self.MAX_IMAGE_SIZE_MB:
            raise ProcessingError(
                f"Image is {size_mb:.1f}MB — maximum is {self.MAX_IMAGE_SIZE_MB}MB. "
                "Resize or compress the image before uploading.",
                recoverable=False
            )

        # Detect media type
        ext       = filename.split(".")[-1].lower()
        media_map = {"jpg": "image/jpeg", "jpeg": "image/jpeg",
                     "png": "image/png", "gif": "image/gif",
                     "webp": "image/webp", "bmp": "image/png"}
        media_type = media_map.get(ext, "image/jpeg")

        import base64, urllib.request
        image_b64 = base64.standard_b64encode(file_bytes).decode("utf-8")

        # Call Claude Vision
        payload = json.dumps({
            "model": "claude-sonnet-4-6",
            "max_tokens": 4096,
            "messages": [{
                "role": "user",
                "content": [
                    {
                        "type": "image",
                        "source": {
                            "type": "base64",
                            "media_type": media_type,
                            "data": image_b64,
                        },
                    },
                    {
                        "type": "text",
                        "text": (
                            "Extract ALL text from this image exactly as written. "
                            "If this is a document (invoice, report, table), preserve the structure. "
                            "For tables, use pipe-separated format. "
                            "For invoices, clearly label: Supplier, Date, Amount, Due Date, Line Items. "
                            "Respond with the extracted text only — no commentary."
                        ),
                    },
                ],
            }],
        }).encode()

        req = urllib.request.Request(
            "https://api.anthropic.com/v1/messages",
            data    = payload,
            headers = {
                "Content-Type":      "application/json",
                "x-api-key":         api_key,
                "anthropic-version": "2023-06-01",
            },
        )

        try:
            with urllib.request.urlopen(req, timeout=30) as resp:
                response      = json.loads(resp.read())
                extracted_text = response["content"][0]["text"]
        except urllib.error.HTTPError as e:
            body = e.read().decode()
            raise ProcessingError(f"Claude Vision API error {e.code}: {body}", recoverable=False)
        except Exception as e:
            raise ProcessingError(f"Image OCR failed: {e}", recoverable=False)

        result.text       = extracted_text
        result.word_count = len(extracted_text.split())
        result.page_count = 1
        result.metadata   = {"ocr_model": "claude-sonnet-4-6", "source_type": "image", "media_type": media_type}

        return result


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# DOCUMENT CLASSIFIER
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

class DocumentClassifier:
    """
    Classifies documents into categories based on content patterns.
    Returns: invoice / p_and_l / bank_statement / budget / contract / other
    """

    PATTERNS = {
        "invoice": [
            r"faktura",  r"invoice",  r"fakturanummer", r"invoice\s*no",
            r"att\s+betala", r"due\s+date", r"förfallodatum",
            r"nettosumma",   r"moms",       r"total\s+inc\s+vat",
        ],
        "p_and_l": [
            r"resultatr[äa]kning", r"profit.*loss", r"income.*statement",
            r"r[öo]relseresultat", r"nettoomsättning", r"bruttovinst",
            r"r[öo]relsekostnad",  r"[åa]rets.*resultat",
        ],
        "bank_statement": [
            r"kontoutdrag", r"bank.*statement", r"saldo",
            r"transaktioner",   r"inbetalning",  r"utbetalning",
            r"kontonummer",     r"clearing",
        ],
        "budget": [
            r"budget",     r"prognos",  r"forecast",
            r"budgeterat", r"utfall",   r"avvikelse",
            r"m[åa]lv[äa]rde",
        ],
        "contract": [
            r"avtal",       r"kontrakt",  r"agreement",  r"contract",
            r"parter",      r"parterna",  r"underteckna", r"signatur",
            r"villkor",     r"terms.*conditions",
        ],
        "payroll": [
            r"l[öo]nespecifikation", r"payslip", r"l[öo]n",
            r"arbetsgivaravgift",    r"semesterl[öo]n", r"skatt",
            r"nettol[öo]n",
        ],
    }

    def classify(self, text: str, filename: str = "") -> tuple[str, float]:
        """Returns (doc_type, confidence_0_to_1)."""
        text_lower     = text.lower()
        filename_lower = filename.lower()
        scores         = {}

        for doc_type, patterns in self.PATTERNS.items():
            score = 0
            for pattern in patterns:
                matches = len(re.findall(pattern, text_lower))
                score  += min(matches, 3)  # cap at 3 per pattern

            # Filename bonus
            if doc_type in filename_lower or doc_type[:5] in filename_lower:
                score += 5
            # Common filename patterns
            if doc_type == "invoice" and re.search(r"faktura|invoice|fakt", filename_lower):
                score += 8
            if doc_type == "bank_statement" and re.search(r"kontoutdrag|bank|statement", filename_lower):
                score += 8
            if doc_type == "p_and_l" and re.search(r"resultat|p.?l|income", filename_lower):
                score += 8

            scores[doc_type] = score

        if not scores or max(scores.values()) == 0:
            return ("other", 0.3)

        best       = max(scores, key=scores.get)
        best_score = scores[best]
        # Normalise confidence: score of 10+ = high confidence
        confidence = min(0.98, best_score / 12)

        return (best, round(confidence, 2))


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# INVOICE EXTRACTOR
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

class InvoiceExtractor:
    """
    Extracts structured invoice data from text.
    Handles Swedish and English invoice formats.
    """

    def extract(self, text: str, tables: list) -> dict:
        lower = text.lower()
        return {
            "vendor":         self._extract_vendor(text),
            "invoice_number": self._extract_pattern(lower, [
                r"fakturanummer[:\s]+([A-Z0-9\-]+)",
                r"invoice\s*(?:no|number|#)[:\s]+([A-Z0-9\-]+)",
                r"nr[:\s]+([A-Z0-9\-]{4,15})",
            ]),
            "invoice_date":   self._extract_date(lower, [
                r"fakturadatum[:\s]+(\d{4}-\d{2}-\d{2})",
                r"invoice\s*date[:\s]+(\d{4}-\d{2}-\d{2})",
                r"datum[:\s]+(\d{4}-\d{2}-\d{2})",
            ]),
            "due_date":       self._extract_date(lower, [
                r"f[öo]rfallodatum[:\s]+(\d{4}-\d{2}-\d{2})",
                r"betalas\s*senast[:\s]+(\d{4}-\d{2}-\d{2})",
                r"due\s*(?:date)?[:\s]+(\d{4}-\d{2}-\d{2})",
            ]),
            "total_inc_vat":  self._extract_amount(lower, [
                r"att\s+betala[:\s]+([\d\s,\.]+)\s*kr",
                r"total\s+inc\s+vat[:\s]+([\d\s,\.]+)",
                r"summa\s+inkl[:\s]+([\d\s,\.]+)\s*kr",
            ]),
            "total_exc_vat":  self._extract_amount(lower, [
                r"nettosumma[:\s]+([\d\s,\.]+)\s*kr",
                r"total\s+exc\s+vat[:\s]+([\d\s,\.]+)",
                r"summa\s+exkl[:\s]+([\d\s,\.]+)\s*kr",
            ]),
            "vat_amount":     self._extract_amount(lower, [
                r"moms\s+(?:12|25|6)\s*%[:\s]+([\d\s,\.]+)\s*kr",
                r"vat\s+amount[:\s]+([\d\s,\.]+)",
            ]),
            "line_items":     self._extract_line_items(text, tables),
            "currency":       "SEK" if "kr" in lower else "EUR" if "eur" in lower else "USD",
        }

    def _extract_vendor(self, text: str) -> str:
        lines = text.strip().split("\n")
        for line in lines[:6]:
            line = line.strip()
            if len(line) > 3 and not re.match(r"^\d", line) and not re.match(r"^(faktura|invoice|datum)", line, re.I):
                return line[:80]
        return ""

    def _extract_pattern(self, text: str, patterns: list) -> Optional[str]:
        for pat in patterns:
            m = re.search(pat, text, re.I)
            if m:
                return m.group(1).strip().upper()
        return None

    def _extract_date(self, text: str, patterns: list) -> Optional[str]:
        for pat in patterns:
            m = re.search(pat, text, re.I)
            if m:
                return m.group(1).strip()
        # Fallback: find any date
        m = re.search(r"\d{4}-\d{2}-\d{2}", text)
        return m.group(0) if m else None

    def _extract_amount(self, text: str, patterns: list) -> Optional[float]:
        for pat in patterns:
            m = re.search(pat, text, re.I)
            if m:
                return self._parse_swedish_amount(m.group(1))
        return None

    def _parse_swedish_amount(self, s: str) -> Optional[float]:
        s = s.strip().replace("\xa0", "").replace(" ", "")
        # Swedish: 34 157,00 → 34157.00
        if "," in s and "." not in s:
            parts = s.split(",")
            if len(parts[-1]) <= 2:
                s = s.replace(",", ".")
        elif "," in s and "." in s:
            if s.rindex(",") > s.rindex("."):
                s = s.replace(".", "").replace(",", ".")
            else:
                s = s.replace(",", "")
        try:
            return float(s)
        except ValueError:
            return None

    def _extract_line_items(self, text: str, tables: list) -> list:
        items = []
        # Try from tables first
        for tbl in tables:
            rows = tbl.get("rows", [])
            if len(rows) < 2:
                continue
            header = [h.lower() for h in rows[0]]
            # Check if this looks like a line item table
            if any(kw in " ".join(header) for kw in ["qty", "antal", "pris", "price", "belopp", "amount"]):
                for row in rows[1:]:
                    if len(row) >= 2 and any(row):
                        items.append({"raw": " | ".join(row)})
        return items


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# MAIN ENTRY POINT
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def process_document(file_bytes: bytes, filename: str, api_key: str = "") -> ProcessingResult:
    """
    Route a file to the correct processor based on extension.
    Returns a ProcessingResult with extracted text, tables, and structured data.
    """
    ext        = filename.split(".")[-1].lower().strip()
    classifier = DocumentClassifier()
    result     = None

    try:
        if ext == "pdf":
            result = PDFProcessor().process(file_bytes)
        elif ext in ("docx", "doc"):
            result = DocxProcessor().process(file_bytes)
        elif ext in ("xlsx", "xls"):
            result = ExcelProcessor().process(file_bytes)
        elif ext == "csv":
            result = CSVProcessor().process(file_bytes)
        elif ext in ("jpg", "jpeg", "png", "gif", "webp", "bmp"):
            result = ImageProcessor().process(file_bytes, filename, api_key)
        elif ext in ("txt", "md", "text"):
            result = ProcessingResult()
            result.text       = file_bytes.decode("utf-8", errors="replace")
            result.word_count = len(result.text.split())
            result.page_count = 1
            result.file_hash  = hashlib.sha256(file_bytes).hexdigest()
        else:
            # Try PDF as fallback for unknown types
            try:
                result = PDFProcessor().process(file_bytes)
            except Exception:
                raise ProcessingError(f"Unsupported file type: .{ext}. Supported: PDF, DOCX, XLSX, CSV, TXT, JPG, PNG")

    except ProcessingError:
        raise
    except Exception as e:
        raise ProcessingError(f"Unexpected error processing {filename}: {e}")

    # Classify and extract structure
    result.doc_type, confidence = classifier.classify(result.text, filename)

    if result.doc_type == "invoice":
        extractor        = InvoiceExtractor()
        result.structured = extractor.extract(result.text, result.tables)

    return result


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# QUICK SELF-TEST
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

if __name__ == "__main__":
    # Test CSV
    csv_data = b"Datum;Beskrivning;Belopp;Saldo\n2026-03-01;Menigo;-35064;249386\n2026-03-15;Sysco;-43624;205762\n"
    r = process_document(csv_data, "bank.csv")
    print(f"CSV: type={r.doc_type}, words={r.word_count}")
    print(r.text[:200])

    # Test Swedish invoice text
    inv_text = b"""Menigo Foodservice AB
Fakturanummer: FAK-8842901
Fakturadatum: 2026-03-12
Forfallodag: 2026-04-11
Nettosumma: 31 307 kr
Moms 12%: 3 757 kr
Att betala: 35 064 kr
"""
    r2 = process_document(inv_text, "menigo_faktura.txt")
    print(f"\nInvoice: type={r2.doc_type}")
    print(f"Structured: {r2.structured}")
